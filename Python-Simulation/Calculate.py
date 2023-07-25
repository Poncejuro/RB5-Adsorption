def Dax_correlation(U):
    #Properties
    Dff=5*10**-14
    rho = 1000 
    eps = 0.13
    mu = 0.000891
    As = 5e6  #https://sci-hub.se/https://doi.org/10.1002/jccs.200200096

    #Surface Area
    d = 4*eps/(As*(1-eps))  #https://sci-hub.se/https://doi.org/10.1016/0923-0467(96)03073-4

    #Reynolds
    Re = U*d*rho/mu

    #Axial dispersion
    Re=d*U*rho/mu/eps   #10^-3 < Re < 10^3
    Sc=mu/Dff/rho
    var=0.72/Re/Sc + 0.52/(1+0.9/Re/Sc)
    Pe=1/var
    DL=2*U*(d/2)/Pe

    return DL, Re


from openpyxl import load_workbook as lb
def Experimental_Data(Name):
    r1='2'
    if Name == 'E1':
        r2='59'
    elif Name == 'E2':
        r2='66'
    elif Name == 'E3':
        r2='65'
    elif Name == 'E4':
        r2='68'
    elif Name == 'E5':
        r2='68'
    elif Name == 'E6':
        r2='56'
    elif Name == 'E7':
        r2='53'
    elif Name == 'E8':
        r2='61'
    elif Name == 'E9':
        r2='55'
    elif Name == 'E10':
        r2='60'
    elif Name == 'E11':
        r2='47'
        
    wb = lb(filename='Datos.xlsx', read_only=True)
    ws = wb[Name]
    # Read the cell values into a list of lists
    time = []
    for row in ws['A'+r1:'A'+r2]:
        for cell in row:
            time.append(cell.value)

    concentration = []
    for row in ws['B'+r1:'B'+r2]:
        for cell in row:
            concentration.append(cell.value)
   
    vars = []
    for row in ws['G1':'G4']:
            for cell in row:
                vars.append(cell.value)
    return time, concentration, vars


